"""
Service module for handling scheduling and data export operations.
"""

import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(results):
    """
    Export schedule results to Excel with formatting.
    
    Args:
        results: List of dictionaries containing schedule data
        
    Returns:
        BytesIO object containing the Excel file
    """
    df = pd.DataFrame(results)
    
    # Create Excel writer
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Schedule', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Schedule']
        
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Format headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Format data cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = center_alignment
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add metadata sheet
        metadata_sheet = workbook.create_sheet('Metadata', 0)
        metadata_sheet['A1'] = 'Invigilation Schedule Report'
        metadata_sheet['A1'].font = Font(bold=True, size=14)
        metadata_sheet['A3'] = 'Generated on:'
        metadata_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        metadata_sheet['A4'] = 'Total Records:'
        metadata_sheet['B4'] = len(results)
        
        # Adjust metadata sheet column width
        metadata_sheet.column_dimensions['A'].width = 20
        metadata_sheet.column_dimensions['B'].width = 30
    
    output.seek(0)
    return output


def generate_statistics(results):
    """
    Generate statistics from schedule results.
    
    Args:
        results: List of dictionaries containing schedule data
        
    Returns:
        Dictionary with statistics
    """
    total_slots = len(results)
    filled_faculty = sum(1 for r in results if r.get('Faculty_1') != '---')
    filled_staff = sum(1 for r in results if r.get('Staff') != '---')
    
    # Count unique assignments
    faculty_assignments = set()
    staff_assignments = set()
    
    for result in results:
        if result.get('Faculty_1') != '---':
            faculty_assignments.add(result['Faculty_1'])
        if result.get('Faculty_2') != '---':
            faculty_assignments.add(result['Faculty_2'])
        if result.get('Staff') != '---':
            staff_assignments.add(result['Staff'])
    
    return {
        'total_slots': total_slots,
        'filled_slots': filled_faculty + filled_staff,
        'faculty_utilization': (filled_faculty / total_slots * 100) if total_slots > 0 else 0,
        'staff_utilization': (filled_staff / total_slots * 100) if total_slots > 0 else 0,
        'unique_faculty': len(faculty_assignments),
        'unique_staff': len(staff_assignments),
    }


def format_schedule_results(results):
    """
    Format schedule results for display.
    
    Args:
        results: List of dictionaries containing schedule data
        
    Returns:
        Formatted results with additional computed fields
    """
    formatted_results = []
    for idx, result in enumerate(results, 1):
        formatted_result = {
            'ID': idx,
            'Date': result.get('Date', 'N/A'),
            'Shift': result.get('Shift', 'N/A'),
            'Room': result.get('Room', 'N/A'),
            'Faculty_1': result.get('Faculty_1', '---'),
            'Faculty_2': result.get('Faculty_2', '---'),
            'Staff': result.get('Staff', '---'),
            'Status': '✓' if result.get('Faculty_1') != '---' else '✗'
        }
        formatted_results.append(formatted_result)
    
    return formatted_results
